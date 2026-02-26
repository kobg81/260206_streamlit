import streamlit as st
import pdfplumber
import re
import pandas as pd
import datetime
import io
from openpyxl.styles import PatternFill

# =====================================================================
# 1. [협약접수확인서] 데이터 추출 함수
# =====================================================================
def extract_receipt(file_stream):
    data = {
        "과제번호": "미추출", 
        "총사업기간": "미추출", 
        "기관정보": [] 
    }
    
    with pdfplumber.open(file_stream) as pdf:
        page = pdf.pages[0] 
        tables = page.extract_tables()
        
        for table in tables:
            df = pd.DataFrame(table).fillna("")
            df = df.applymap(lambda x: str(x).replace('\n', ' ').strip())
            
            for i, row in df.iterrows():
                row_str = "".join(row).replace(" ", "")
                if "과제번호" in row_str:
                    m = re.search(r'(RS-\d{4}-\d{8}|\b\d{8}\b)', row_str)
                    if m: data["과제번호"] = m.group(1)
                if "총사업기간" in row_str:
                    dates = re.findall(r'\d{4}-\d{2}-\d{2}', row_str)
                    if len(dates) >= 2: data["총사업기간"] = f"{dates[0]}~{dates[1]}"
            
            if "참여역할" in df.values:
                headers = list(df.iloc[0])
                if "기관명" in headers and "사업자번호" in headers:
                    r_idx = headers.index("참여역할")
                    n_idx = headers.index("기관명")
                    b_idx = headers.index("사업자번호")
                    for _, r in df.iloc[1:].iterrows():
                        if r[n_idx]:
                            data["기관정보"].append({
                                "역할": r[r_idx], "기관명": r[n_idx], "사업자번호": r[b_idx]
                            })
    return data

# =====================================================================
# 2. [협약서] 데이터 추출 및 검증 함수
# =====================================================================
def extract_agreement(file_stream, r_data):
    info = {
        "과제번호": "미추출", "과제명": "미추출", "전체기간": "미추출", 
        "세부기간정보": [], 
        "검증결과": {}
    }
    
    with pdfplumber.open(file_stream) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"
        clean_text = re.sub(r'\s+', ' ', full_text)

        no_match = re.search(r'(?<!총괄)연구개발과제번호\s*[:\s]*(RS-\d{4}-\d{8})', clean_text)
        if no_match: info["과제번호"] = no_match.group(1)

        target_no = info["과제번호"] if info["과제번호"] != "미추출" else r_data["과제번호"]
        title_pattern = re.search(r'따라\s+(?:연구개발과제\s*)?(.*?)(?:\s*연구개발과제\s*)?\(\s*연구개발과제번호', clean_text)
        
        if title_pattern:
            info["과제명"] = title_pattern.group(1).strip()
        else:
            for page in pdf.pages:
                for table in page.extract_tables():
                    for row in table:
                        row_str = " ".join([str(c) for c in row if c])
                        if "연구개발과제명" in row_str:
                            parts = [str(c).replace('\n', ' ').strip() for c in row if c and len(str(c)) > 15 and "연구개발과제명" not in str(c)]
                            if parts: 
                                info["과제명"] = parts[0]
                                break
                    if info["과제명"] != "미추출": break
                if info["과제명"] != "미추출": break

        if info["과제명"] == "미추출":
            info["과제명"] = "미추출 (※ 확인필요: 제1조(목적) 텍스트 양식 상이)"

        period_pattern = re.search(r'전체\s*([\d\.\s]+[-~]\s*[\d\.\s]+)', clean_text)
        if not period_pattern: 
            for page in pdf.pages:
                for table in page.extract_tables():
                    for row in table:
                        row_str = "".join([str(c) for c in row if c]).replace(" ", "")
                        if "전체" in row_str:
                            m = re.search(r'(\d{4}[\.\d]*[-~]\d{4}[\.\d]*)', row_str)
                            if m: 
                                info["전체기간"] = m.group(1)
                                break
        else:
            info["전체기간"] = period_pattern.group(1).split('(')[0].strip().rstrip('.')

        current_stage = "1단계" 
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    row_str_spaced = " ".join([str(c).replace('\n', ' ') for c in row if c])
                    
                    stage_match = re.search(r'(\d+)\s*단계', row_str_spaced)
                    if stage_match:
                        current_stage = f"{stage_match.group(1)}단계"
                    
                    year_match = re.search(r'(\d+)\s*년차\s*[\(\[]\s*(\d{4}[-\.]\d{2}(?:[-\.]\d{2})?)\s*[~-]\s*(\d{4}[-\.]\d{2}(?:[-\.]\d{2})?)\s*[\)\]]', row_str_spaced)
                    
                    if year_match:
                        y_num = year_match.group(1)
                        start_date = year_match.group(2)
                        end_date = year_match.group(3)
                        
                        unique_key = f"{current_stage}_{y_num}년차"
                        if not any(d['key'] == unique_key for d in info["세부기간정보"]):
                            info["세부기간정보"].append({
                                "key": unique_key,
                                "단계": current_stage,
                                "년차": f"{y_num}년차",
                                "기간": f"{start_date} ~ {end_date}"
                            })

    if info["세부기간정보"]:
        info["세부기간정보"] = sorted(
            info["세부기간정보"], 
            key=lambda x: (int(re.search(r'\d+', x['단계']).group()), int(re.search(r'\d+', x['년차']).group()))
        )

    r_no = r_data["과제번호"]
    r_period = r_data.get("총사업기간", "")
    r_p_clean = re.sub(r'[^0-9]', '', r_period)
    a_p_clean = re.sub(r'[^0-9]', '', info["전체기간"])
    
    period_match = False
    if len(r_p_clean) >= 12 and len(a_p_clean) >= 12: 
        period_match = (r_p_clean[:6] == a_p_clean[:6]) 

    info["검증결과"] = {
        "과제번호": "✅ 일치" if info["과제번호"] == r_no else "❌ 불일치",
        "사업기간": "✅ 일치" if period_match else "❌ 불일치"
    }
    return info

# =====================================================================
# 3. [연구비분담 추출] 엑셀 데이터 추출 함수
# =====================================================================
def get_clean_int_str(val):
    if pd.notnull(val) and str(val).strip() != "":
        try:
            return str(int(float(val)))
        except ValueError:
            return str(val).strip()
    return ""

def extract_fund_excel(file_stream):
    fund_data = {}
    try:
        df = pd.read_excel(file_stream, header=1)
        
        for idx, row in df.iterrows():
            org_name = str(row.iloc[3]).strip()
            if org_name == "nan" or not org_name: continue
            
            stage = get_clean_int_str(row.iloc[0])
            year_count = get_clean_int_str(row.iloc[1])
            year = get_clean_int_str(row.iloc[2])
            
            gov_cash = int(row.iloc[4]) * 1000 if pd.notnull(row.iloc[4]) else 0
            inst_cash = int(row.iloc[6]) * 1000 if pd.notnull(row.iloc[6]) else 0
            inst_kind = int(row.iloc[7]) * 1000 if pd.notnull(row.iloc[7]) else 0
            
            if org_name not in fund_data:
                fund_data[org_name] = []
            
            fund_data[org_name].append({
                "당해년도": year,
                "단계_년차": f"{stage}단계 {year_count}년차",
                "정부지원금(현금)": f"{gov_cash:,}",
                "기관부담금(현금)": f"{inst_cash:,}",
                "기관부담금(현물)": f"{inst_kind:,}"
            })
            
    except Exception as e:
        st.error(f"엑셀 파일 읽기 오류: {e}")
        return {}
        
    return fund_data

# =====================================================================
# 4. [엑셀파일생성] 다운로드용 엑셀 데이터 생성 함수
# =====================================================================
def generate_excel_file(agreement_data, sheet2_data, current_year, receipt_data):
    output = io.BytesIO()
    proj_no = agreement_data.get("과제번호", "미추출")
    
    # 주관연구개발기관 및 사업자번호 탐색
    lead_org_name = ""
    lead_biz_no = ""
    if receipt_data and "기관정보" in receipt_data:
        for org in receipt_data["기관정보"]:
            if "주관" in org.get("역할", "") or "총괄" in org.get("역할", ""):
                lead_org_name = org.get("기관명", "")
                lead_biz_no = org.get("사업자번호", "")
                break

    # 당해년도(current_year)에 해당하는 세부 기간 파싱
    current_stage = "-"
    current_year_count = "-"
    start_date_str = "-"
    end_date_str = "-"
    
    for period_info in agreement_data.get("세부기간정보", []):
        if current_year in period_info["기간"]:
            current_stage = period_info["단계"] # 예: 1단계
            current_year_count = period_info["년차"].replace("년차", "차년도") # 예: 1차년도
            
            dates = period_info["기간"].split("~")
            if len(dates) == 2:
                s_match = re.search(r'(\d{4})[-\.](\d{2})', dates[0])
                e_match = re.search(r'(\d{4})[-\.](\d{2})', dates[1])
                if s_match: start_date_str = f"{s_match.group(1)}.{s_match.group(2)}"
                if e_match: end_date_str = f"{e_match.group(1)}.{e_match.group(2)}"
            break
    
    # 1열~10열 데이터 프레임 구성
    df_sheet1 = pd.DataFrame([{
        "GI_ACC_NO": "",  # <--- 1열에 빈 값으로 추가
        "1열 과제번호": proj_no,
        "2열 과제명": agreement_data.get("과제명", ""),
        "3열 총사업기간": agreement_data.get("전체기간", ""),
        "4열 주관연구개발기관명": lead_org_name,
        "5열 사업자번호": lead_biz_no,
        "6열 당해년도": f"{current_year}년",
        "7열 당해년도 단계": current_stage,
        "8열 당해년도 년차": current_year_count,
        "9열 년차 시작일": start_date_str,
        "10열 년차 종료일": end_date_str
    }])
    
    df_sheet2 = pd.DataFrame(sheet2_data)
    
    sheet1_name = f"{proj_no}_과제기본정보"[:31]
    sheet2_name = f"{proj_no}_참여기관정보"[:31]
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_sheet1.to_excel(writer, index=False, sheet_name=sheet1_name)
        if not df_sheet2.empty:
            df_sheet2.to_excel(writer, index=False, sheet_name=sheet2_name)
        
        workbook = writer.book
        blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        
        # Sheet1 스타일링
        ws1 = workbook[sheet1_name]
        ws1.column_dimensions['A'].width = 15 # GI_ACC_NO
        ws1.column_dimensions['B'].width = 20 # 1열 과제번호
        ws1.column_dimensions['C'].width = 50 # 2열 과제명 (넓게)
        ws1.column_dimensions['D'].width = 30 # 3열 총사업기간
        ws1.column_dimensions['E'].width = 30 # 4열 기관명
        ws1.column_dimensions['F'].width = 20 # 5열 사업자번호
        ws1.column_dimensions['G'].width = 15
        ws1.column_dimensions['H'].width = 15
        ws1.column_dimensions['I'].width = 15
        ws1.column_dimensions['J'].width = 15
        
        if not df_sheet2.empty:
            ws2 = workbook[sheet2_name]
            ws2.column_dimensions['A'].width = 15 # GI_ACC_NO
            ws2.column_dimensions['B'].width = 25 # 1열 개발기관명
            ws2.column_dimensions['C'].width = 15 # 2열 사업자번호
            ws2.column_dimensions['D'].width = 25 # 3열 역할
            ws2.column_dimensions['E'].width = 25
            ws2.column_dimensions['F'].width = 25
            ws2.column_dimensions['G'].width = 25

        for ws in workbook.worksheets:
            for row in ws.iter_rows(min_row=2): 
                is_target = any("한국전자기술연구원" in str(cell.value) for cell in row if cell.value)
                if is_target:
                    for cell in row:
                        cell.fill = blue_fill

    output.seek(0)
    return output
# =====================================================================
# 5. Streamlit UI
# =====================================================================
st.set_page_config(page_title="연구관리 통합 검증기", layout="wide")
st.title("🛡️ 협약 데이터 및 당해년도 연구비 통합 검증")

col1, col2 = st.columns(2)
with col1: 
    f_agreement = st.file_uploader("1. 협약서 (PDF) - 왼쪽", type=['pdf'])
with col2: 
    f_receipt = st.file_uploader("2. 협약접수확인서 (PDF) - 오른쪽", type=['pdf'])

st.write("---")
f_fund = st.file_uploader("3. 연구비분담표 (Excel) - 하단", type=['xlsx', 'xls'])

if f_receipt and f_agreement and f_fund:
    receipt_data = extract_receipt(f_receipt)
    agreement_data = extract_agreement(f_agreement, receipt_data)
    fund_excel_data = extract_fund_excel(f_fund)
    
    current_year = str(datetime.datetime.now().year)
    
    st.divider()
    st.subheader("📊 기본 정보 검증 결과")
    v1, v2 = st.columns(2)
    v1.metric("0. 연구개발과제번호", agreement_data["과제번호"], delta=agreement_data["검증결과"]["과제번호"])
    v2.metric("1. 총 사업기간", agreement_data["전체기간"], delta=agreement_data["검증결과"]["사업기간"])
    
    if "미추출" in agreement_data['과제명']:
        st.warning(f"**과제명 (협약서 기준):** {agreement_data['과제명']}")
    else:
        st.info(f"**과제명 (협약서 기준):** {agreement_data['과제명']}") 
    
    st.divider()

    sheet2_export_data = []
    all_orgs = receipt_data["기관정보"]
    sorted_orgs = sorted(all_orgs, key=lambda x: ("총괄" not in x["역할"], "주관" not in x["역할"])) 

    for org in sorted_orgs:
        role_title = "총괄주관연구개발기관" if "총괄" in org["역할"] else ("주관연구개발기관" if "주관" in org["역할"] else "공동연구개발기관")
        biz_no = org.get("사업자번호", "-")
        
        matched_funds = []
        for fname, f_list in fund_excel_data.items():
            if org['기관명'].replace(" ", "") in fname.replace(" ", "") or fname.replace(" ", "") in org['기관명'].replace(" ", ""):
                matched_funds.extend(f_list)
        
        current_year_funds = [f for f in matched_funds if f["당해년도"] == current_year]
        
        if current_year_funds:
            for fund in current_year_funds:
                sheet2_export_data.append({
			    "GI_ACC_NO 계정번호": "",  # <--- 1열에 빈 값으로 추가
                            "GI_ORG 개발기관명": org['기관명'],
                            "GT_BUSSNO 사업자번호": biz_no,
                            "3열 주관연구개발기관 또는 공동연구개발기관": role_title,
                            "4열 당해년도": fund['당해년도'],
                            "GI_GOVFUND 정부지원금(현금)": fund['정부지원금(현금)'],
                            "GI_PRIVCASH 연구개발기관 부담금(현금)": fund['기관부담금(현금)'],
                            "GI_PRIVINK 연구개발기관 부담금(현물)": fund['기관부담금(현물)']
                })
        else:
            sheet2_export_data.append({
		"GI_ACC_NO 계정번호": "",  # <--- 1열에 빈 값으로 추가
                "GI_ORG 공동연구개발기관명": org['기관명'],
                "GT_BUSSNO 사업자번호": biz_no,
                "3열 주관연구개발기관 또는 공동연구개발기관": role_title,
                "4열 당해년도": current_year,
                "GI_GOVFUND 정부지원금(현금)": "-",
                "GI_PRIVCASH 연구개발기관 부담금(현금)": "-",
                "GI_PRIVINK 연구개발기관 부담금(현물)": "-"
            })

    excel_file_bytes = generate_excel_file(agreement_data, sheet2_export_data, current_year, receipt_data)
    excel_file_name = f"{agreement_data['과제번호']}_{current_year}_과제정보.xlsx"
    
    st.download_button(
        label="📥 과제정보 다운로드 (Excel)",
        data=excel_file_bytes,
        file_name=excel_file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, 
        type="primary" 
    )
    
    st.divider()

    tab1, tab2 = st.tabs(["💰 당해년도 연구비 현황", "📅 전체 단계/년차 일정"])
    
    with tab1:
        st.info(f"💡 **'당해년도'는 프로그램 실행 날짜 기준입니다.** (현재 설정 연도: {current_year}년)")
        
        for org in sorted_orgs:
            role_title = "총괄주관연구개발기관" if "총괄" in org["역할"] else ("주관연구개발기관" if "주관" in org["역할"] else "공동연구개발기관")
            
            with st.expander(f"{role_title} : {org['기관명']} ({org['사업자번호']})", expanded=True):
                matched_funds = []
                for fname, f_list in fund_excel_data.items():
                    if org['기관명'].replace(" ", "") in fname.replace(" ", "") or fname.replace(" ", "") in org['기관명'].replace(" ", ""):
                        matched_funds.extend(f_list)
                
                current_year_funds = [f for f in matched_funds if f["당해년도"] == current_year]
                
                if current_year_funds:
                    for fund in current_year_funds:
                        st.markdown(f"**당해년도({fund['당해년도']}) | {fund['단계_년차']}**")
                        st.write(f"▸ 정부지원금(현금): `{fund['정부지원금(현금)']}` / 연구개발기관 부담금(현금): `{fund['기관부담금(현금)']}` / 연구개발기관 부담금(현물): `{fund['기관부담금(현물)']}`")
                else:
                    st.warning(f"엑셀 파일에서 당해년도({current_year}년) 연구비 정보를 찾을 수 없습니다. (과제 기간을 확인해주세요)")

    with tab2:
        if agreement_data["세부기간정보"]:
            df_period = pd.DataFrame(agreement_data["세부기간정보"])
            st.dataframe(df_period[["단계", "년차", "기간"]], use_container_width=True, hide_index=True)
        else:
            st.warning("⚠️ 단계/년차별 기간 정보를 찾을 수 없습니다.")

elif not (f_receipt and f_agreement and f_fund):
    st.info("👆 위 3개의 파일을 모두 업로드하면 분석 결과가 표시됩니다.")